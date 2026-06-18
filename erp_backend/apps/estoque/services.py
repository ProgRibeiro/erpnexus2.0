import csv
import io
import re
import unicodedata
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import Count
from django.db.utils import OperationalError, ProgrammingError
from openpyxl import load_workbook

from apps.configuracoes.models import get_empresa_configurada

from .models import CategoriaProduto, MotorInteligenciaConhecimento, Produto, Servico


class MemoriaMotorInteligente:
    STOPWORDS = {
        "para", "com", "sem", "uma", "uns", "das", "dos", "por", "que", "nao", "não",
        "quando", "entao", "então", "deve", "devo", "usar", "sugira", "sugerir",
        "cliente", "servico", "serviço", "produto", "orcamento", "orçamento",
    }

    def buscar(self, texto, escopo=None, limite=5, incrementar=False):
        tokens = self._tokens(texto)
        if not tokens:
            return []

        escopos = [MotorInteligenciaConhecimento.Escopo.GERAL]
        if escopo:
            escopos.append(escopo)

        qs = (
            MotorInteligenciaConhecimento.objects
            .select_related("produto", "servico")
            .filter(ativo=True, escopo__in=escopos)
            .order_by("-atualizado_em")[:400]
        )
        ranqueados = []
        try:
            conhecimentos = list(qs)
        except (OperationalError, ProgrammingError):
            return []

        for item in conhecimentos:
            texto_item = " ".join([
                item.titulo,
                item.entrada,
                item.resposta,
                " ".join(item.termos or []),
                getattr(item.produto, "nome", "") or "",
                getattr(item.servico, "nome", "") or "",
            ])
            item_tokens = self._tokens(texto_item)
            intersecao = tokens.intersection(item_tokens)
            if not intersecao:
                continue
            score = len(intersecao) * 10 + min(item.confianca, 100) / 10 + min(item.vezes_usado, 20)
            ranqueados.append((score, item, sorted(intersecao)))

        ranqueados.sort(key=lambda registro: registro[0], reverse=True)
        conhecimentos = [self._serializar(item, score, termos) for score, item, termos in ranqueados[:limite]]
        if incrementar and conhecimentos:
            ids = [item["id"] for item in conhecimentos]
            for conhecimento in MotorInteligenciaConhecimento.objects.filter(id__in=ids):
                conhecimento.vezes_usado += 1
                conhecimento.save(update_fields=["vezes_usado"])
        return conhecimentos

    def ensinar_por_chat(self, mensagem, usuario=None, contexto=None):
        contexto = contexto or {}
        dados = self._extrair_ensino(mensagem, contexto)
        conhecimento = MotorInteligenciaConhecimento.objects.create(
            titulo=dados["titulo"],
            escopo=dados["escopo"],
            tipo=dados["tipo"],
            entrada=dados["entrada"],
            resposta=dados["resposta"],
            termos=sorted(self._tokens(" ".join([dados["titulo"], dados["entrada"], dados["resposta"]]))),
            payload=dados["payload"],
            produto_id=dados.get("produto"),
            servico_id=dados.get("servico"),
            confianca=dados["confianca"],
            origem=MotorInteligenciaConhecimento.Origem.CHAT,
            status_revisao=MotorInteligenciaConhecimento.StatusRevisao.APROVADO,
            criado_por=usuario if getattr(usuario, "is_authenticated", False) else None,
        )
        return conhecimento

    def responder_chat(self, mensagem, usuario=None, contexto=None):
        texto = str(mensagem or "").strip()
        contexto = contexto or {}
        if self._parece_ensino(texto):
            conhecimento = self.ensinar_por_chat(texto, usuario=usuario, contexto=contexto)
            return {
                "acao": "aprendido",
                "resposta": f"Aprendi: {conhecimento.titulo}. Vou usar isso em {conhecimento.get_escopo_display().lower()}.",
                "conhecimento": self._serializar(conhecimento, 100, conhecimento.termos),
                "sugestoes": self.buscar(texto, escopo=conhecimento.escopo, limite=3),
            }

        escopo = contexto.get("escopo") or self._inferir_escopo(texto)
        encontrados = self.buscar(texto, escopo=escopo, limite=5, incrementar=True)
        if encontrados:
            principal = encontrados[0]
            return {
                "acao": "resposta",
                "resposta": principal["resposta"],
                "conhecimento": principal,
                "sugestoes": encontrados,
            }

        return {
            "acao": "sem_memoria",
            "resposta": "Ainda não tenho uma regra específica para isso. Você pode me ensinar escrevendo: 'Quando acontecer X, sugira Y'.",
            "conhecimento": None,
            "sugestoes": [],
        }

    def aplicar_em_orcamento(self, texto):
        conhecimentos = self.buscar(texto, escopo=MotorInteligenciaConhecimento.Escopo.ORCAMENTO, limite=6, incrementar=True)
        itens = []
        tipo_servico = None
        prioridade = None
        for conhecimento in conhecimentos:
            payload = conhecimento.get("payload") or {}
            tipo_servico = tipo_servico or payload.get("tipo_servico")
            prioridade = prioridade or payload.get("prioridade")
            if conhecimento.get("servico_id"):
                try:
                    servico = Servico.objects.get(id=conhecimento["servico_id"], ativo=True)
                    itens.append(("servico", servico, conhecimento))
                except Servico.DoesNotExist:
                    pass
            if conhecimento.get("produto_id"):
                try:
                    produto = Produto.objects.get(id=conhecimento["produto_id"], ativo=True)
                    itens.append(("produto", produto, conhecimento))
                except Produto.DoesNotExist:
                    pass
        return {"conhecimentos": conhecimentos, "itens": itens, "tipo_servico": tipo_servico, "prioridade": prioridade}

    def aprender_com_ordem(self, ordem, usuario=None):
        if not ordem or not getattr(ordem, "pk", None):
            return []
        if ordem.status not in ["concluida", "faturada"]:
            return []

        itens = list(ordem.itens.select_related("produto", "servico").all())
        if not itens:
            return []

        entrada = self._entrada_da_ordem(ordem, itens)
        termos = sorted(self._tokens(entrada))
        payload_base = self._payload_ordem(ordem, itens)
        aprendidos = []

        for item in itens:
            produto = item.produto if item.origem_tipo == "produto" else None
            servico = item.servico if item.origem_tipo == "servico" else None
            if not produto and not servico and not item.descricao:
                continue

            alvo = servico or produto
            titulo_alvo = getattr(alvo, "nome", "") or item.descricao[:80]
            tipo = (
                MotorInteligenciaConhecimento.Tipo.PRECO
                if item.valor_unitario
                else MotorInteligenciaConhecimento.Tipo.REGRA
            )
            resposta = self._resposta_ordem(ordem, item, produto=produto, servico=servico)
            payload = {
                **payload_base,
                "item_origem_id": item.id,
                "valor_unitario": str(item.valor_unitario or 0),
                "quantidade_media": str(item.quantidade or 1),
            }
            filtros = {
                "os_origem": ordem,
                "produto": produto,
                "servico": servico,
                "entrada": entrada[:1000],
            }
            conhecimento, _ = MotorInteligenciaConhecimento.objects.update_or_create(
                **filtros,
                defaults={
                    "titulo": f"{titulo_alvo} em {ordem.get_tipo_servico_display()}",
                    "escopo": MotorInteligenciaConhecimento.Escopo.ORCAMENTO,
                    "tipo": tipo,
                    "resposta": resposta,
                    "termos": termos,
                    "payload": payload,
                    "confianca": self._confianca_aprendizado(ordem, item, produto, servico),
                    "ativo": True,
                    "origem": MotorInteligenciaConhecimento.Origem.OS_CONCLUIDA,
                    "status_revisao": MotorInteligenciaConhecimento.StatusRevisao.PENDENTE,
                    "criado_por": usuario if getattr(usuario, "is_authenticated", False) else None,
                },
            )
            aprendidos.append(conhecimento)
        return aprendidos

    def aprender_com_ordens_concluidas(self, limite=20, usuario=None):
        from apps.ordens.models import OrdemServico

        ordens = (
            OrdemServico.objects
            .filter(status__in=[OrdemServico.Status.CONCLUIDA, OrdemServico.Status.FATURADA])
            .prefetch_related("itens__produto", "itens__servico")
            .order_by("-atualizado_em")[: int(limite or 20)]
        )
        aprendidos = []
        for ordem in ordens:
            aprendidos.extend(self.aprender_com_ordem(ordem, usuario=usuario))
        return aprendidos

    def painel_memoria(self):
        qs = MotorInteligenciaConhecimento.objects.all()
        por_status = qs.values("status_revisao").annotate(total=Count("id"))
        por_origem = qs.values("origem").annotate(total=Count("id"))
        return {
            "total": qs.count(),
            "ativos": qs.filter(ativo=True).count(),
            "pendentes": qs.filter(status_revisao=MotorInteligenciaConhecimento.StatusRevisao.PENDENTE).count(),
            "aprovados": qs.filter(status_revisao=MotorInteligenciaConhecimento.StatusRevisao.APROVADO).count(),
            "rejeitados": qs.filter(status_revisao=MotorInteligenciaConhecimento.StatusRevisao.REJEITADO).count(),
            "por_status": {item["status_revisao"]: item["total"] for item in por_status},
            "por_origem": {item["origem"]: item["total"] for item in por_origem},
        }

    def _extrair_ensino(self, mensagem, contexto):
        texto = str(mensagem or "").strip()
        escopo = contexto.get("escopo") or self._inferir_escopo(texto)
        tipo = contexto.get("tipo") or MotorInteligenciaConhecimento.Tipo.REGRA
        entrada = texto
        resposta = texto

        match = re.search(r"quando\s+(.+?)(?:,\s*|\s+)(?:sugira|use|usar|considere|responda|entao|então)\s+(.+)", texto, re.IGNORECASE)
        if match:
            entrada = match.group(1).strip()
            resposta = match.group(2).strip()

        payload = {}
        tipo_servico = self._inferir_tipo_servico(texto)
        if tipo_servico:
            payload["tipo_servico"] = tipo_servico
        if any(chave in self._normalizar(texto) for chave in ["urgente", "emergencia", "emergência", "parado"]):
            payload["prioridade"] = "urgente"

        servico = self._melhor_servico(texto)
        produto = self._melhor_produto(texto)

        titulo = contexto.get("titulo") or entrada[:90] or "Conhecimento ensinado"
        return {
            "titulo": titulo,
            "escopo": escopo,
            "tipo": tipo,
            "entrada": entrada,
            "resposta": resposta,
            "payload": payload,
            "produto": produto.id if produto else None,
            "servico": servico.id if servico else None,
            "confianca": int(contexto.get("confianca") or 85),
        }

    def _parece_ensino(self, texto):
        normalizado = self._normalizar(texto)
        return normalizado.startswith(("aprenda", "ensine", "ensinar")) or "quando " in normalizado and any(
            chave in normalizado for chave in ["sugira", "use", "usar", "considere", "responda", "entao", "então"]
        )

    def _inferir_escopo(self, texto):
        normalizado = self._normalizar(texto)
        if any(chave in normalizado for chave in ["orcamento", "orçamento", "os ", "cliente pediu", "sugira"]):
            return MotorInteligenciaConhecimento.Escopo.ORCAMENTO
        if any(chave in normalizado for chave in ["catalogo", "catálogo", "produto", "servico", "serviço", "estoque"]):
            return MotorInteligenciaConhecimento.Escopo.CATALOGO
        return MotorInteligenciaConhecimento.Escopo.GERAL

    def _inferir_tipo_servico(self, texto):
        normalizado = self._normalizar(texto)
        mapa = {
            "hvac": ["hvac", "split", "ar condicionado", "condensadora", "evaporadora"],
            "refrigeracao": ["refrigeracao", "refrigeração", "camara", "câmara", "freezer"],
            "eletrica": ["eletrica", "elétrica", "disjuntor", "quadro", "tomada", "curto"],
            "civil": ["civil", "pintura", "parede", "gesso", "piso"],
            "instalacao": ["instalacao", "instalação", "instalar"],
            "manutencao": ["manutencao", "manutenção", "preventiva", "corretiva"],
        }
        for tipo, termos in mapa.items():
            if any(termo in normalizado for termo in termos):
                return tipo
        return None

    def _melhor_servico(self, texto):
        tokens = self._tokens(texto)
        melhor = (0, None)
        for servico in Servico.objects.filter(ativo=True)[:300]:
            score = len(tokens.intersection(self._tokens(" ".join([servico.nome, servico.descricao, servico.categoria]))))
            if score > melhor[0]:
                melhor = (score, servico)
        return melhor[1] if melhor[0] >= 2 else None

    def _melhor_produto(self, texto):
        tokens = self._tokens(texto)
        melhor = (0, None)
        for produto in Produto.objects.select_related("categoria").filter(ativo=True)[:500]:
            score = len(tokens.intersection(self._tokens(" ".join([
                produto.nome,
                produto.descricao,
                produto.categoria.nome if produto.categoria else "",
            ]))))
            if score > melhor[0]:
                melhor = (score, produto)
        return melhor[1] if melhor[0] >= 2 else None

    def _serializar(self, item, score, termos):
        return {
            "id": item.id,
            "titulo": item.titulo,
            "escopo": item.escopo,
            "tipo": item.tipo,
            "entrada": item.entrada,
            "resposta": item.resposta,
            "payload": item.payload,
            "produto_id": item.produto_id,
            "produto_nome": item.produto.nome if item.produto else "",
            "servico_id": item.servico_id,
            "servico_nome": item.servico.nome if item.servico else "",
            "confianca": item.confianca,
            "vezes_usado": item.vezes_usado,
            "origem": item.origem,
            "status_revisao": item.status_revisao,
            "os_origem_id": item.os_origem_id,
            "score": round(float(score), 2),
            "termos_match": list(termos or []),
        }

    def _entrada_da_ordem(self, ordem, itens):
        partes = [
            ordem.descricao_servico,
            ordem.observacoes_tecnicas,
            ordem.equipamento_marca,
            ordem.equipamento_modelo,
            ordem.equipamento_serie,
            " ".join(item.descricao for item in itens if item.descricao),
        ]
        texto = " ".join(parte for parte in partes if parte)
        return texto[:1000] or f"OS {ordem.numero or ordem.id} {ordem.get_tipo_servico_display()}"

    def _payload_ordem(self, ordem, itens):
        produtos = [
            {"id": item.produto_id, "nome": item.produto.nome, "quantidade": str(item.quantidade)}
            for item in itens
            if item.produto_id and item.produto
        ]
        servicos = [
            {"id": item.servico_id, "nome": item.servico.nome, "quantidade": str(item.quantidade)}
            for item in itens
            if item.servico_id and item.servico
        ]
        checklist = [
            resposta.item.texto
            for resposta in ordem.respostas_checklist.select_related("item").all()[:12]
            if resposta.valor_bool is True or resposta.valor_texto or resposta.valor_numero is not None
        ]
        return {
            "tipo_servico": ordem.tipo_servico,
            "prioridade": ordem.prioridade,
            "cliente_id": ordem.cliente_id,
            "cliente_nome": getattr(ordem.cliente, "nome", ""),
            "os_numero": ordem.numero,
            "valor_total_orcado": str(ordem.valor_total_orcado or 0),
            "valor_final_faturado": str(ordem.valor_final_faturado or 0),
            "produtos_usados": produtos,
            "servicos_executados": servicos,
            "checklist_sugerido": checklist,
        }

    def _resposta_ordem(self, ordem, item, produto=None, servico=None):
        alvo = getattr(servico or produto, "nome", "") or item.descricao
        partes = [f"Em casos parecidos com {ordem.get_tipo_servico_display()}, considerar {alvo}."]
        if item.valor_unitario:
            partes.append(f"Valor unitário praticado: R$ {Decimal(item.valor_unitario).quantize(Decimal('0.01'))}.")
        if produto:
            partes.append("Verificar disponibilidade no estoque antes de aprovar.")
        if ordem.prioridade in ["alta", "urgente"]:
            partes.append(f"Prioridade histórica: {ordem.get_prioridade_display()}.")
        return " ".join(partes)

    def _confianca_aprendizado(self, ordem, item, produto, servico):
        score = 58
        if servico:
            score += 14
        if produto:
            score += 10
        if item.valor_unitario:
            score += 8
        if ordem.valor_final_faturado or ordem.valor_total_orcado:
            score += 5
        if ordem.respostas_checklist.exists():
            score += 5
        return min(score, 92)

    def _tokens(self, texto):
        return {
            token
            for token in re.findall(r"[a-z0-9]{3,}", self._normalizar(texto))
            if token not in self.STOPWORDS
        }

    def _normalizar(self, value):
        texto = unicodedata.normalize("NFKD", str(value or "").lower())
        texto = "".join(char for char in texto if not unicodedata.combining(char))
        return re.sub(r"\s+", " ", texto).strip()


class MotorCatalogoInteligente:
    COLUNAS_ALIASES = {
        "tipo": {"tipo", "classe", "origem", "natureza", "tipo item", "tipo do item"},
        "nome": {"nome", "item", "produto", "material", "servico", "serviço", "nome item", "nome do item"},
        "descricao": {"descricao", "descrição", "detalhe", "detalhes", "descricao_completa", "especificacao", "especificação", "especificacoes", "especificações", "especificacao tecnica", "especificação técnica", "caracteristicas", "características"},
        "observacoes": {"observacao", "observação", "observacoes", "observações", "obs", "comentario", "comentário"},
        "categoria": {"categoria", "grupo", "familia", "família", "tipo_servico", "linha"},
        "subcategoria": {"subcategoria", "sub categoria", "subgrupo", "sub grupo", "classe material", "classe"},
        "unidade": {"unidade", "un", "und", "medida", "unidade_medida", "unidade de medida", "u.m.", "um"},
        "custo": {"custo", "preco_custo", "preço_custo", "compra", "preco_compra", "preço_compra", "valor custo", "valor compra"},
        "venda": {"venda", "preco_venda", "preço_venda", "preco", "preço", "valor", "valor_venda", "preco ref", "preço ref", "preco referencia", "preço referência", "preco de referencia", "preço de referência", "preco ref r", "preço ref r"},
        "markup": {"markup", "margem", "margem_percentual"},
        "estoque_minimo": {"estoque_minimo", "estoque mínimo", "minimo", "mínimo"},
        "localizacao": {"localizacao", "localização", "prateleira", "local"},
        "tributacao": {"tributacao", "tributação", "imposto", "tributo"},
        "codigo_lc116": {"codigo_lc116", "código_lc116", "lc116", "codigo_servico"},
        "marca": {"marca", "fabricante", "marca referencia", "marca referência", "marca_ref"},
    }

    PRODUTO_KEYWORDS = {
        "capacitor", "rele", "relé", "filtro", "gas", "gás", "cabo", "fio", "disjuntor",
        "tubo", "mangueira", "parafuso", "sensor", "placa", "controle", "compressor",
        "motor", "contator", "bobina", "valvula", "válvula", "carregador", "wallbox",
    }
    SERVICO_KEYWORDS = {
        "limpeza", "higienizacao", "higienização", "instalacao", "instalação", "manutencao",
        "manutenção", "diagnostico", "diagnóstico", "mao de obra", "mão de obra",
        "visita", "troca", "reparo", "correcao", "correção", "preventiva", "corretiva",
    }

    CATEGORIAS_SERVICO = {
        "hvac": {"hvac", "ar", "split", "condensadora", "evaporadora", "climatizacao", "climatização"},
        "refrigeracao": {"refrigeracao", "refrigeração", "freezer", "camara", "câmara", "geladeira"},
        "eletrica": {"eletrica", "elétrica", "quadro", "disjuntor", "tomada", "cabo", "fio"},
        "civil": {"civil", "parede", "pintura", "gesso", "forro", "piso", "alvenaria"},
        "manutencao": {"manutencao", "manutenção", "preventiva", "corretiva", "revisao", "revisão"},
        "instalacao": {"instalacao", "instalação", "instalar", "montagem"},
    }

    UNIDADES_PRODUTO = {choice[0] for choice in Produto.UnidadeMedida.choices}
    UNIDADES_SERVICO = {choice[0] for choice in Servico.UnidadeMedida.choices}
    SHEET_SKIP_KEYWORDS = {"resumo", "summary", "dashboard", "totais", "total", "grafico", "gráfico"}

    def analisar(self, texto="", arquivo=None, markup_padrao=None, despesas_padrao=None):
        linhas = self._extrair_linhas(texto=texto, arquivo=arquivo)
        fiscal = self._dados_fiscais()
        markup = self._decimal(markup_padrao, Decimal("35.00"))
        despesas = self._decimal(despesas_padrao, Decimal("8.00"))

        itens = []
        avisos = []
        for indice, linha in enumerate(linhas, start=1):
            item = self._normalizar_linha(linha, indice, fiscal, markup, despesas)
            if item["nome"]:
                itens.append(item)
            else:
                avisos.append(f"Linha {indice} ignorada por falta de nome/descrição.")

        resumo = {
            "total_linhas": len(linhas),
            "produtos": sum(1 for item in itens if item["tipo"] == "produto"),
            "servicos": sum(1 for item in itens if item["tipo"] == "servico"),
            "valor_total_venda": str(sum((Decimal(item["preco_venda"]) for item in itens), Decimal("0.00")).quantize(Decimal("0.01"))),
            "fiscal": fiscal,
            "modelos_suportados": [
                "Nome/Descrição/Categoria/Unidade/Custo/Venda",
                "Código/Categoria/Subcategoria/Item/Especificação Técnica/Unidade/Preço Ref./Marca",
                "CSV com ; , tab ou |",
                "Texto livre com partes separadas por ; ou |",
            ],
        }
        return {"itens": itens, "resumo": resumo, "avisos": avisos}

    def criar(self, itens):
        resultado = {"produtos_criados": 0, "produtos_atualizados": 0, "servicos_criados": 0, "servicos_atualizados": 0, "erros": []}
        criados = []

        for indice, item in enumerate(itens, start=1):
            try:
                with transaction.atomic():
                    if item.get("tipo") == "servico":
                        obj, created = self._criar_servico(item)
                        resultado["servicos_criados" if created else "servicos_atualizados"] += 1
                    else:
                        obj, created = self._criar_produto(item)
                        resultado["produtos_criados" if created else "produtos_atualizados"] += 1
                criados.append({"tipo": item.get("tipo"), "id": obj.id, "nome": obj.nome, "created": created})
            except Exception as exc:
                resultado["erros"].append(f"Linha {indice}: {exc}")

        resultado["itens"] = criados
        return resultado

    def _criar_produto(self, item):
        categoria_nome = item.get("categoria") or "Geral"
        categoria, _ = CategoriaProduto.objects.get_or_create(
            nome=categoria_nome,
            defaults={"descricao": "Categoria criada pelo motor inteligente de catálogo.", "ativo": True},
        )
        defaults = {
            "descricao": item.get("descricao", ""),
            "categoria": categoria,
            "unidade_medida": item.get("unidade_medida") if item.get("unidade_medida") in self.UNIDADES_PRODUTO else Produto.UnidadeMedida.UN,
            "preco_custo": Decimal(str(item.get("preco_custo") or 0)),
            "preco_venda": Decimal(str(item.get("preco_venda") or 0)),
            "markup_percentual": Decimal(str(item.get("markup_percentual") or 0)),
            "aliquota_impostos_percentual": Decimal(str(item.get("aliquota_impostos_percentual") or 0)),
            "despesas_operacionais_percentual": Decimal(str(item.get("despesas_operacionais_percentual") or 0)),
            "preco_manual": True,
            "estoque_minimo": Decimal(str(item.get("estoque_minimo") or 0)),
            "localizacao": item.get("localizacao", ""),
            "ativo": True,
        }
        codigo = item.get("codigo")
        if codigo:
            obj, created = Produto.objects.update_or_create(codigo=codigo, defaults={"nome": item["nome"], **defaults})
        else:
            obj, created = Produto.objects.update_or_create(nome=item["nome"], defaults=defaults)
        return obj, created

    def _criar_servico(self, item):
        categoria = item.get("categoria") if item.get("categoria") in dict(Servico.Categoria.choices) else self._inferir_categoria_servico(item["nome"])
        defaults = {
            "descricao": item.get("descricao", ""),
            "categoria": categoria,
            "unidade_medida": item.get("unidade_medida") if item.get("unidade_medida") in self.UNIDADES_SERVICO else Servico.UnidadeMedida.UNI,
            "preco_padrao": Decimal(str(item.get("preco_venda") or 0)),
            "tributacao": item.get("tributacao") if item.get("tributacao") in dict(Servico.Tributacao.choices) else Servico.Tributacao.ISS,
            "codigo_lc116": item.get("codigo_lc116", ""),
            "ativo": True,
        }
        codigo = item.get("codigo")
        if codigo:
            obj, created = Servico.objects.update_or_create(codigo=codigo, defaults={"nome": item["nome"], **defaults})
        else:
            obj, created = Servico.objects.update_or_create(nome=item["nome"], defaults=defaults)
        return obj, created

    def _normalizar_linha(self, linha, indice, fiscal, markup_padrao, despesas_padrao):
        nome = self._limpar(linha.get("nome") or linha.get("descricao") or linha.get("subcategoria") or "")
        descricao_partes = [
            linha.get("descricao"),
            linha.get("subcategoria"),
            linha.get("marca") and f"Marca referência: {linha.get('marca')}",
            linha.get("observacoes"),
        ]
        descricao = self._limpar(" | ".join(str(parte).strip() for parte in descricao_partes if str(parte or "").strip()))
        texto = self._normalizar(" ".join([
            nome,
            descricao,
            str(linha.get("categoria") or ""),
            str(linha.get("subcategoria") or ""),
            str(linha.get("tipo") or ""),
            str(linha.get("codigo") or ""),
        ]))
        memoria = MemoriaMotorInteligente().buscar(
            texto,
            escopo=MotorInteligenciaConhecimento.Escopo.CATALOGO,
            limite=3,
            incrementar=True,
        )
        tipo_informado = linha.get("tipo")
        tipo = self._inferir_tipo(tipo_informado, texto)
        if not self._normalizar(tipo_informado) and (linha.get("marca") or linha.get("subcategoria")) and linha.get("venda"):
            tipo = "produto"
        tipo = (memoria[0].get("payload") or {}).get("tipo") or tipo if memoria else tipo
        if memoria and memoria[0].get("servico_id"):
            tipo = "servico"
        if memoria and memoria[0].get("produto_id"):
            tipo = "produto"
        categoria = self._inferir_categoria(linha.get("categoria"), texto, tipo)
        categoria = (memoria[0].get("payload") or {}).get("categoria") or categoria if memoria else categoria
        unidade = self._inferir_unidade(linha.get("unidade"), tipo)
        custo = self._decimal(linha.get("custo"), Decimal("0.00"))
        venda_informada = self._decimal(linha.get("venda"), None)
        if custo == 0 and venda_informada is not None:
            custo = venda_informada
        markup = self._decimal(linha.get("markup"), markup_padrao)
        despesas = despesas_padrao if tipo == "produto" else Decimal("0.00")
        aliquota = self._aliquota_item(tipo, fiscal, linha.get("tributacao"))
        venda = venda_informada if venda_informada is not None else self._calcular_venda(custo, markup, aliquota, despesas)
        if venda == 0 and custo == 0:
            venda = self._preco_padrao_por_texto(texto, tipo)
        if memoria and (memoria[0].get("payload") or {}).get("preco_venda"):
            venda = self._decimal(memoria[0]["payload"].get("preco_venda"), venda)

        return {
            "linha": indice,
            "tipo": tipo,
            "codigo": self._limpar(linha.get("codigo") or ""),
            "nome": nome[:255],
            "descricao": descricao,
            "categoria": categoria,
            "unidade_medida": unidade,
            "preco_custo": str(custo.quantize(Decimal("0.01"))),
            "preco_venda": str(venda.quantize(Decimal("0.01"))),
            "markup_percentual": str(markup.quantize(Decimal("0.01"))),
            "aliquota_impostos_percentual": str(aliquota.quantize(Decimal("0.01"))),
            "despesas_operacionais_percentual": str(despesas.quantize(Decimal("0.01"))),
            "estoque_minimo": str(self._decimal(linha.get("estoque_minimo"), Decimal("0")).quantize(Decimal("0.01"))),
            "localizacao": self._limpar(linha.get("localizacao") or ""),
            "tributacao": "iss" if tipo == "servico" else self._limpar(linha.get("tributacao") or "icms"),
            "codigo_lc116": self._limpar(linha.get("codigo_lc116") or ""),
            "motivo": self._motivo(tipo, venda_informada, fiscal, memoria),
            "memoria_aplicada": memoria,
        }

    def _extrair_linhas(self, texto="", arquivo=None):
        linhas = []
        if arquivo:
            nome = str(getattr(arquivo, "name", "")).lower()
            if nome.endswith((".xlsx", ".xlsm", ".xls")):
                linhas.extend(self._linhas_xlsx(arquivo))
            else:
                conteudo = arquivo.read().decode("utf-8-sig", errors="ignore")
                linhas.extend(self._linhas_csv_ou_texto(conteudo))
        if texto:
            linhas.extend(self._linhas_csv_ou_texto(texto))
        return linhas

    def _linhas_xlsx(self, arquivo):
        if hasattr(arquivo, "seek"):
            arquivo.seek(0)
        wb = load_workbook(io.BytesIO(arquivo.read()), data_only=True, read_only=True)
        linhas = []
        for ws in wb.worksheets:
            nome_aba = self._normalizar(ws.title)
            if any(keyword in nome_aba for keyword in self.SHEET_SKIP_KEYWORDS):
                continue
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header_index = self._detectar_linha_cabecalho(rows)
            if header_index is None:
                continue
            headers = [self._map_header(value, index) for index, value in enumerate(rows[header_index])]
            for row in rows[header_index + 1:]:
                if not any(value not in (None, "") for value in row):
                    continue
                item = {
                    headers[index]: value
                    for index, value in enumerate(row)
                    if index < len(headers) and headers[index] and not headers[index].startswith("coluna_")
                }
                if self._linha_tem_conteudo_catalogo(item):
                    item["_aba"] = ws.title
                    linhas.append(item)
        return linhas

    def _detectar_linha_cabecalho(self, rows):
        melhor = (0, None)
        for index, row in enumerate(rows[:20]):
            headers = [self._map_header(value, col_index) for col_index, value in enumerate(row)]
            score = self._score_headers(headers)
            if score > melhor[0]:
                melhor = (score, index)
        return melhor[1] if melhor[0] >= 2 else 0

    def _score_headers(self, headers):
        reconhecidos = set(headers)
        score = 0
        for campo in {"nome", "descricao", "categoria", "subcategoria", "unidade", "custo", "venda", "codigo", "marca"}:
            if campo in reconhecidos:
                score += 1
        if "nome" in reconhecidos and ("venda" in reconhecidos or "custo" in reconhecidos):
            score += 2
        return score

    def _linha_tem_conteudo_catalogo(self, item):
        return bool(self._limpar(item.get("nome") or item.get("descricao") or item.get("codigo")))

    def _linhas_csv_ou_texto(self, texto):
        amostra = str(texto or "").strip()
        if not amostra:
            return []
        delimitador = ";" if amostra.count(";") >= amostra.count(",") else ","
        if "\t" in amostra:
            delimitador = "\t"
        primeira = amostra.splitlines()[0]
        if delimitador in primeira and self._parece_cabecalho(primeira, delimitador):
            reader = csv.DictReader(io.StringIO(amostra), delimiter=delimitador)
            return [{self._map_header(k, i): v for i, (k, v) in enumerate(row.items())} for row in reader]

        linhas = []
        for raw in amostra.splitlines():
            linha = raw.strip(" -•\t")
            if not linha:
                continue
            partes = [parte.strip() for parte in re.split(r"\s*[;|]\s*", linha) if parte.strip()]
            if len(partes) >= 2:
                linhas.append(self._partes_para_linha(partes))
            else:
                linhas.append({"nome": linha})
        return linhas

    def _parece_cabecalho(self, primeira_linha, delimitador):
        celulas = [self._normalizar(celula) for celula in primeira_linha.split(delimitador)]
        campos_reconhecidos = {
            self._map_header(celula, index)
            for index, celula in enumerate(celulas)
            if celula
        }
        campos_base = {"nome", "descricao", "tipo", "categoria", "subcategoria", "unidade", "custo", "venda", "codigo"}
        return len(campos_reconhecidos.intersection(campos_base)) >= 2

    def _partes_para_linha(self, partes):
        linha = {"nome": partes[0]}
        for parte in partes[1:]:
            chave_valor = re.split(r"\s*[:=]\s*", parte, maxsplit=1)
            if len(chave_valor) == 2:
                linha[self._map_header(chave_valor[0], 0)] = chave_valor[1]
            elif self._normalizar(parte) in {"produto", "material", "peca", "peça", "servico", "serviço"}:
                linha["tipo"] = parte
            elif re.search(r"\d", parte):
                linha.setdefault("venda", parte)
            else:
                linha.setdefault("categoria", parte)
        return linha

    def _map_header(self, value, index):
        texto = self._normalizar(value)
        texto = re.sub(r"\([^)]*\)", "", texto).strip()
        texto = re.sub(r"[^a-z0-9]+", " ", texto).strip()
        for destino, aliases in self.COLUNAS_ALIASES.items():
            aliases_normalizados = {
                re.sub(r"[^a-z0-9]+", " ", self._normalizar(alias)).strip()
                for alias in aliases
            }
            if texto in aliases_normalizados:
                return destino
        if texto in {"codigo", "sku", "cod", "cod item", "referencia", "referência"}:
            return "codigo"
        return f"coluna_{index + 1}" if not texto else texto

    def _dados_fiscais(self):
        try:
            empresa = get_empresa_configurada()
            produto = sum(Decimal(str(v or 0)) for v in [
                empresa.aliquota_pis_padrao,
                empresa.aliquota_cofins_padrao,
                empresa.aliquota_irpj_padrao,
                empresa.aliquota_csll_padrao,
            ])
            return {
                "regime_tributario": empresa.regime_tributario or "simples_nacional",
                "aliquota_servico": str(Decimal(str(empresa.aliquota_issqn_padrao or 0)).quantize(Decimal("0.01"))),
                "aliquota_produto": str(produto.quantize(Decimal("0.01"))),
            }
        except (OperationalError, ProgrammingError):
            return {
                "regime_tributario": "nao_configurado",
                "aliquota_servico": "0.00",
                "aliquota_produto": "0.00",
            }

    def _aliquota_item(self, tipo, fiscal, tributacao):
        if tipo == "servico":
            return Decimal(fiscal["aliquota_servico"])
        return Decimal(fiscal["aliquota_produto"])

    def _calcular_venda(self, custo, markup, aliquota, despesas):
        return custo + (custo * (markup + aliquota + despesas) / Decimal("100"))

    def _inferir_tipo(self, informado, texto):
        valor = self._normalizar(informado)
        if valor in {"servico", "serviço", "mao de obra", "mão de obra"}:
            return "servico"
        if valor in {"produto", "material", "peca", "peça"}:
            return "produto"
        if any(chave in texto for chave in self.SERVICO_KEYWORDS):
            return "servico"
        if any(chave in texto for chave in self.PRODUTO_KEYWORDS):
            return "produto"
        return "produto"

    def _inferir_categoria(self, informado, texto, tipo):
        informado_norm = self._normalizar(informado)
        if tipo == "servico":
            if informado_norm in dict(Servico.Categoria.choices):
                return informado_norm
            return self._inferir_categoria_servico(texto)
        return self._limpar(informado) or self._categoria_produto_por_texto(texto)

    def _inferir_categoria_servico(self, texto):
        texto_norm = self._normalizar(texto)
        for categoria, palavras in self.CATEGORIAS_SERVICO.items():
            if any(self._normalizar(palavra) in texto_norm for palavra in palavras):
                return categoria
        return "manutencao"

    def _categoria_produto_por_texto(self, texto):
        if any(chave in texto for chave in ["ar", "split", "capacitor", "rele", "relé", "gas", "gás"]):
            return "Ar condicionado / HVAC"
        if any(chave in texto for chave in ["cabo", "fio", "disjuntor", "tomada", "lampada", "lâmpada", "led", "eletroduto", "interruptor", "quadro", "dr", "dps"]):
            return "Elétrica"
        return "Geral"

    def _inferir_unidade(self, informado, tipo):
        unidade = self._normalizar(informado)
        unidade = unidade.replace(".", "")
        mapa = {
            "un": "un",
            "und": "un",
            "unid": "un",
            "unidade": "un",
            "pc": "un",
            "peca": "un",
            "peça": "un",
            "metro": "m",
            "metros": "m",
            "m": "m",
            "m2": "m2",
            "m²": "m2",
            "metro quadrado": "m2",
            "kg": "kg",
            "quilograma": "kg",
            "l": "litro",
            "lt": "litro",
            "litro": "litro",
            "par": "par",
            "caixa": "caixa",
            "cx": "caixa",
            "hora": "hora",
            "hr": "hora",
            "dia": "dia",
            "lote": "lote",
        }
        unidade = mapa.get(unidade, unidade)
        validas = self.UNIDADES_SERVICO if tipo == "servico" else self.UNIDADES_PRODUTO
        if unidade in validas:
            return unidade
        return "uni" if tipo == "servico" else "un"

    def _preco_padrao_por_texto(self, texto, tipo):
        if tipo == "servico":
            if "instal" in texto:
                return Decimal("850.00")
            if "limpeza" in texto or "higien" in texto:
                return Decimal("280.00")
            if "diagn" in texto or "visita" in texto:
                return Decimal("250.00")
            return Decimal("350.00")
        return Decimal("0.00")

    def _motivo(self, tipo, venda_informada, fiscal, memoria=None):
        if memoria:
            return f"Memória aplicada: {memoria[0]['titulo']}."
        if venda_informada is not None:
            return "Preço de venda informado na entrada; impostos fiscais mantidos para referência."
        if tipo == "servico":
            return f"Preço calculado com ISS padrão de {fiscal['aliquota_servico']}% e margem informada/padrão."
        return f"Preço calculado com carga fiscal padrão de {fiscal['aliquota_produto']}%, despesas e margem."

    def _decimal(self, value, default):
        if value in (None, ""):
            return default
        texto = str(value).strip().replace("R$", "").replace("%", "").replace(" ", "")
        if "," in texto and "." in texto:
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", ".")
        try:
            return Decimal(texto)
        except (InvalidOperation, ValueError):
            return default

    def _limpar(self, value):
        return str(value or "").strip()

    def _normalizar(self, value):
        texto = unicodedata.normalize("NFKD", str(value or "").lower())
        texto = "".join(char for char in texto if not unicodedata.combining(char))
        return re.sub(r"\s+", " ", texto).strip()
