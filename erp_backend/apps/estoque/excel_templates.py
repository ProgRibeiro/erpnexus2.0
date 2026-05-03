from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO


class ExcelTemplateGenerator:
    @staticmethod
    def gerar_template_clientes():
        """Gera template Excel para importação de clientes"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        headers = [
            "Nome Fantasia*",
            "Razão Social",
            "CNPJ",
            "Telefone",
            "Email",
            "Segmento",
            "Status",
            "Cidade",
            "UF",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1B4F8A", end_color="1B4F8A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 10

        exemplo = ["Empresa XYZ", "EMPRESA XYZ LTDA", "12.345.678/0001-90", "(11) 99999-9999", "contato@xyz.com", "comercio", "ativo", "São Paulo", "SP"]
        for col, valor in enumerate(exemplo, 1):
            ws.cell(row=2, column=col, value=valor)

        arquivo = BytesIO()
        wb.save(arquivo)
        arquivo.seek(0)
        return arquivo

    @staticmethod
    def gerar_template_servicos():
        """Gera template Excel para importação de serviços"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Serviços"

        headers = ["Nome*", "Descrição", "Categoria*", "Preço*", "Tributação", "LC116", "Unidade"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1B4F8A", end_color="1B4F8A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 12

        exemplo = ["Serviço HVAC", "Manutenção de ar condicionado", "hvac", 150.00, "iss", "01.01", "hora"]
        for col, valor in enumerate(exemplo, 1):
            ws.cell(row=2, column=col, value=valor)

        arquivo = BytesIO()
        wb.save(arquivo)
        arquivo.seek(0)
        return arquivo

    @staticmethod
    def gerar_template_produtos():
        """Gera template Excel para importação de produtos"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Produtos"

        headers = ["Nome*", "Descrição", "Categoria", "Unidade", "Custo*", "Venda*", "Mín.", "Localização"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1B4F8A", end_color="1B4F8A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 15

        exemplo = ["Tubo de cobre", "Tubo de cobre para instalação", "Materiais", "un", 25.50, 45.00, 10, "Prateleira A1"]
        for col, valor in enumerate(exemplo, 1):
            ws.cell(row=2, column=col, value=valor)

        arquivo = BytesIO()
        wb.save(arquivo)
        arquivo.seek(0)
        return arquivo
