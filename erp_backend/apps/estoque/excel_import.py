from io import BytesIO
from openpyxl import load_workbook
from apps.clientes.models import Cliente


class ExcelImporter:
    @staticmethod
    def importar_clientes(arquivo):
        """Importa clientes a partir de arquivo Excel"""
        try:
            wb = load_workbook(BytesIO(arquivo.read()))
            ws = wb.active

            sucesso = 0
            falhas = 0
            erros = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row[0]:  # nome é obrigatório
                        continue

                    nome_fantasia = row[0] if row[0] else ""
                    razao_social = row[1] if len(row) > 1 and row[1] else nome_fantasia
                    cnpj = row[2] if len(row) > 2 and row[2] else ""
                    telefone = row[3] if len(row) > 3 and row[3] else ""
                    email = row[4] if len(row) > 4 and row[4] else ""
                    segmento = row[5] if len(row) > 5 and row[5] else ""
                    status = row[6] if len(row) > 6 and row[6] else "ativo"
                    cidade = row[7] if len(row) > 7 and row[7] else ""
                    uf = row[8] if len(row) > 8 and row[8] else ""

                    cliente, created = Cliente.objects.get_or_create(
                        cnpj_cpf=cnpj if cnpj else None,
                        defaults={
                            "nome": nome_fantasia,
                            "nome_fantasia": nome_fantasia,
                            "razao_social": razao_social,
                            "telefone": telefone,
                            "email": email,
                            "segmento": segmento if segmento in dict(Cliente.Segmento.choices) else None,
                            "status": status if status in ["ativo", "inativo"] else "ativo",
                            "cidade": cidade,
                            "uf": uf,
                        }
                    )
                    sucesso += 1
                except Exception as e:
                    falhas += 1
                    erros.append(f"Linha {row_idx}: {str(e)}")

            return {"sucesso": sucesso, "falhas": falhas, "erros": erros}
        except Exception as e:
            return {"sucesso": 0, "falhas": 1, "erros": [str(e)]}

    @staticmethod
    def importar_servicos(arquivo):
        """Importa serviços a partir de planilhas em múltiplos formatos."""
        return ExcelImporter._importar_catalogo(arquivo, tipo_forcado="servico")

    @staticmethod
    def importar_produtos(arquivo):
        """Importa produtos a partir de planilhas em múltiplos formatos."""
        return ExcelImporter._importar_catalogo(arquivo, tipo_forcado="produto")

    @staticmethod
    def _importar_catalogo(arquivo, tipo_forcado=None):
        """Usa o motor inteligente de catálogo para aceitar vários layouts de planilha."""
        try:
            from apps.estoque.services import MotorCatalogoInteligente

            motor = MotorCatalogoInteligente()
            analise = motor.analisar(arquivo=arquivo)
            itens = analise.get("itens", [])
            if tipo_forcado:
                itens = [{**item, "tipo": tipo_forcado} for item in itens]
            resultado = motor.criar(itens)
            criados = resultado.get("produtos_criados", 0) + resultado.get("servicos_criados", 0)
            atualizados = resultado.get("produtos_atualizados", 0) + resultado.get("servicos_atualizados", 0)
            erros = resultado.get("erros", [])
            return {
                "sucesso": criados + atualizados,
                "falhas": len(erros),
                "erros": erros,
                "resumo": analise.get("resumo", {}),
                "avisos": analise.get("avisos", []),
                "criados": criados,
                "atualizados": atualizados,
            }
        except Exception as e:
            return {"sucesso": 0, "falhas": 1, "erros": [str(e)]}
