import csv
import io
import re
import unicodedata
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from django.db import transaction
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.clientes.models import Cliente
from apps.estoque.models import Produto, Servico
from apps.financeiro.models import Lancamento
from apps.ordens.models import OrdemServico


def _xlsx_response(workbook, filename):
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _bold_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _normalizar_header(value):
    texto = unicodedata.normalize("NFKD", str(value or ""))
    texto = "".join(char for char in texto if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "_", texto.strip().lower()).strip("_")


def _texto(value):
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decimal(value, default="0"):
    if value in (None, ""):
        return Decimal(default)
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    texto = _texto(value).replace("R$", "").replace("%", "").replace(" ", "")
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return Decimal(texto)
    except InvalidOperation as exc:
        raise ValueError(f'valor numérico inválido: "{value}"') from exc


def _linhas_arquivo(arquivo, obrigatorias=()):
    extensao = Path(arquivo.name or "").suffix.lower()
    if extensao in {".xlsx", ".xlsm"}:
        workbook = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
    elif extensao in {".csv", ".txt", ".tsv"}:
        conteudo = arquivo.read()
        if not conteudo:
            raise ValueError("O arquivo está vazio.")
        try:
            texto = conteudo.decode("utf-8-sig")
        except UnicodeDecodeError:
            texto = conteudo.decode("cp1252")
        amostra = texto[:8192]
        delimitador_padrao = "\t" if extensao == ".tsv" else ";"
        try:
            dialect = csv.Sniffer().sniff(amostra, delimiters=";,\t|")
            rows = csv.reader(io.StringIO(texto), dialect)
        except csv.Error:
            rows = csv.reader(io.StringIO(texto), delimiter=delimitador_padrao)
    else:
        raise ValueError("Formato não suportado. Envie um arquivo XLSX, XLSM, CSV, TSV ou TXT.")

    iterator = iter(rows)
    try:
        headers = [_normalizar_header(value) for value in next(iterator)]
    except StopIteration as exc:
        raise ValueError("O arquivo está vazio.") from exc
    if not any(headers):
        raise ValueError("A primeira linha deve conter os nomes das colunas.")
    if len(headers) != len(set(filter(None, headers))):
        raise ValueError("A planilha possui nomes de colunas repetidos.")
    faltantes = [coluna for coluna in obrigatorias if coluna not in headers]
    if faltantes:
        raise ValueError(f"Coluna obrigatória ausente: {', '.join(faltantes)}.")

    for row_num, row in enumerate(iterator, start=2):
        if not any(value not in (None, "") for value in row):
            continue
        yield row_num, dict(zip(headers, row))


def _bool(value, default=True):
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "sim", "s", "yes", "ativo"}


def _resultado(criados=0, atualizados=0, ignorados=0, erros=None):
    erros = erros or []
    return Response(
        {
            "criados": criados,
            "atualizados": atualizados,
            "ignorados": ignorados,
            "erros": erros,
            "sucesso": criados + atualizados,
            "falhas": len(erros),
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def exportar_clientes(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    _bold_header(ws, ["id", "nome", "cnpj_cpf", "email", "telefone", "status", "segmento", "criado_em"])
    for c in Cliente.objects.all().order_by("nome"):
        ws.append([
            c.pk,
            c.nome,
            c.cnpj_cpf,
            c.email,
            c.telefone,
            c.status,
            c.segmento,
            c.criado_em.strftime("%Y-%m-%d %H:%M") if c.criado_em else "",
        ])
    return _xlsx_response(wb, f"clientes_{date.today()}.xlsx")


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def importar_clientes(request):
    arquivo = request.FILES.get("arquivo")
    if not arquivo:
        return Response({"erro": "Arquivo não enviado."}, status=400)

    criados = 0
    atualizados = 0
    ignorados = 0
    erros = []

    try:
        linhas = _linhas_arquivo(arquivo, ["nome"])
        for row_num, data in linhas:
            try:
                nome = _texto(data.get("nome"))
                if not nome:
                    ignorados += 1
                    continue
                cnpj_cpf = _texto(data.get("cnpj_cpf") or data.get("cpf_cnpj"))
                status_cliente = _texto(data.get("status")).lower() or Cliente.Status.ATIVO
                if status_cliente not in Cliente.Status.values:
                    raise ValueError(f'status inválido: "{status_cliente}"')
                defaults = {
                    "nome": nome,
                    "email": _texto(data.get("email")),
                    "telefone": _texto(data.get("telefone")),
                    "segmento": _texto(data.get("segmento")),
                    "status": status_cliente,
                }
                with transaction.atomic():
                    if cnpj_cpf:
                        _, created = Cliente.objects.update_or_create(
                            cnpj_cpf=cnpj_cpf,
                            defaults=defaults,
                        )
                        criados += int(created)
                        atualizados += int(not created)
                    else:
                        Cliente.objects.create(cnpj_cpf="", **defaults)
                        criados += 1
            except Exception as e:
                erros.append(f"Linha {row_num}: {e}")
    except Exception as e:
        return Response({"erro": f"Erro ao processar arquivo: {e}"}, status=400)

    return _resultado(criados=criados, atualizados=atualizados, ignorados=ignorados, erros=erros)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def exportar_produtos(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produtos"
    _bold_header(ws, ["id", "codigo", "nome", "descricao", "unidade_medida", "preco_custo", "preco_venda", "estoque_minimo", "ativo"])
    for p in Produto.objects.all().order_by("nome"):
        ws.append([
            p.pk,
            p.codigo or "",
            p.nome,
            p.descricao,
            p.unidade_medida,
            float(p.preco_custo),
            float(p.preco_venda),
            float(p.estoque_minimo),
            p.ativo,
        ])
    return _xlsx_response(wb, f"produtos_{date.today()}.xlsx")


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def importar_produtos(request):
    arquivo = request.FILES.get("arquivo")
    if not arquivo:
        return Response({"erro": "Arquivo não enviado."}, status=400)

    criados = 0
    atualizados = 0
    erros = []

    try:
        linhas = _linhas_arquivo(arquivo, ["nome"])
        for row_num, data in linhas:
            try:
                nome = _texto(data.get("nome"))
                codigo = _texto(data.get("codigo")) or None
                if not nome:
                    continue
                unidade = _texto(data.get("unidade_medida")).lower() or Produto.UnidadeMedida.UN
                if unidade not in Produto.UnidadeMedida.values:
                    raise ValueError(f'unidade_medida inválida: "{unidade}"')
                preco_venda_informado = data.get("preco_venda") not in (None, "")
                defaults = {
                    "nome": nome,
                    "descricao": _texto(data.get("descricao")),
                    "unidade_medida": unidade,
                    "preco_custo": _decimal(data.get("preco_custo")),
                    "preco_venda": _decimal(data.get("preco_venda")),
                    "preco_manual": preco_venda_informado,
                    "estoque_minimo": _decimal(data.get("estoque_minimo")),
                    "ativo": _bool(data.get("ativo"), True),
                }
                with transaction.atomic():
                    if codigo:
                        _, created = Produto.objects.update_or_create(codigo=codigo, defaults=defaults)
                        if created:
                            criados += 1
                        else:
                            atualizados += 1
                    else:
                        Produto.objects.create(**defaults)
                        criados += 1
            except Exception as e:
                erros.append(f"Linha {row_num}: {e}")
    except Exception as e:
        return Response({"erro": f"Erro ao processar arquivo: {e}"}, status=400)

    return _resultado(criados=criados, atualizados=atualizados, erros=erros)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def exportar_servicos(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Servicos"
    _bold_header(ws, ["id", "codigo", "nome", "descricao", "categoria", "unidade_medida", "preco_padrao", "tributacao", "codigo_lc116", "ativo"])
    for s in Servico.objects.all().order_by("categoria", "nome"):
        ws.append([
            s.pk,
            s.codigo or "",
            s.nome,
            s.descricao,
            s.categoria,
            s.unidade_medida,
            float(s.preco_padrao),
            s.tributacao,
            s.codigo_lc116,
            s.ativo,
        ])
    return _xlsx_response(wb, f"servicos_{date.today()}.xlsx")


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def importar_servicos(request):
    arquivo = request.FILES.get("arquivo")
    if not arquivo:
        return Response({"erro": "Arquivo não enviado."}, status=400)

    criados = 0
    atualizados = 0
    erros = []

    try:
        linhas = _linhas_arquivo(arquivo, ["nome"])
        for row_num, data in linhas:
            try:
                nome = _texto(data.get("nome"))
                codigo = _texto(data.get("codigo")) or None
                if not nome:
                    continue
                categoria = _texto(data.get("categoria")).lower() or Servico.Categoria.HVAC
                unidade = _texto(data.get("unidade_medida")).lower() or Servico.UnidadeMedida.UNI
                tributacao = _texto(data.get("tributacao")).lower() or Servico.Tributacao.ISS
                if categoria not in Servico.Categoria.values:
                    raise ValueError(f'categoria inválida: "{categoria}"')
                if unidade not in Servico.UnidadeMedida.values:
                    raise ValueError(f'unidade_medida inválida: "{unidade}"')
                if tributacao not in Servico.Tributacao.values:
                    raise ValueError(f'tributação inválida: "{tributacao}"')
                defaults = {
                    "nome": nome,
                    "descricao": _texto(data.get("descricao")),
                    "categoria": categoria,
                    "unidade_medida": unidade,
                    "preco_padrao": _decimal(data.get("preco_padrao") or data.get("preco")),
                    "tributacao": tributacao,
                    "codigo_lc116": _texto(data.get("codigo_lc116")),
                    "ativo": _bool(data.get("ativo"), True),
                }
                with transaction.atomic():
                    if codigo:
                        _, created = Servico.objects.update_or_create(codigo=codigo, defaults=defaults)
                        criados += int(created)
                        atualizados += int(not created)
                    else:
                        Servico.objects.create(**defaults)
                        criados += 1
            except Exception as e:
                erros.append(f"Linha {row_num}: {e}")
    except Exception as e:
        return Response({"erro": f"Erro ao processar arquivo: {e}"}, status=400)

    return _resultado(criados=criados, atualizados=atualizados, erros=erros)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def exportar_ordens(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ordens de Servico"
    _bold_header(ws, ["numero", "status", "cliente", "tipo_servico", "prioridade", "valor_total_orcado", "valor_final_faturado", "tecnico", "data_agendada", "criado_em"])
    for o in OrdemServico.objects.select_related("cliente", "tecnico_responsavel").order_by("-criado_em")[:5000]:
        ws.append([
            o.numero or "",
            o.status,
            str(o.cliente) if o.cliente else "",
            o.tipo_servico,
            o.prioridade,
            float(o.valor_total_orcado or 0),
            float(o.valor_final_faturado or 0),
            str(o.tecnico_responsavel) if o.tecnico_responsavel else "",
            str(o.data_agendada) if o.data_agendada else "",
            o.criado_em.strftime("%Y-%m-%d %H:%M") if o.criado_em else "",
        ])
    return _xlsx_response(wb, f"ordens_{date.today()}.xlsx")


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def exportar_financeiro(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lancamentos"
    _bold_header(ws, ["id", "tipo", "descricao", "valor", "data_competencia", "data_vencimento", "data_pagamento", "status", "categoria", "conta", "fornecedor_cliente"])
    for l in Lancamento.objects.select_related("categoria", "conta_bancaria").order_by("-data_vencimento")[:10000]:
        ws.append([
            l.pk,
            l.tipo,
            l.descricao,
            float(l.valor),
            str(l.data_competencia),
            str(l.data_vencimento),
            str(l.data_pagamento) if l.data_pagamento else "",
            l.status,
            str(l.categoria) if l.categoria else "",
            str(l.conta_bancaria) if l.conta_bancaria else "",
            l.fornecedor_cliente,
        ])
    return _xlsx_response(wb, f"financeiro_{date.today()}.xlsx")
