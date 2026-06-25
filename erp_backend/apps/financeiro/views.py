import csv
import io
from datetime import date, datetime, timedelta
from decimal import Decimal

from django.db.models import Case, DecimalField, F, Sum, Value, When
from django.db.models.functions import Coalesce, TruncMonth
from django.utils import timezone
from rest_framework import status, viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.ordens.models import OrdemServico

from .models import AnexoLancamento, CategoriaFinanceira, ContaBancaria, Lancamento, TransferenciaEntreConta
from .serializers import (
    AnexoLancamentoSerializer,
    CategoriaFinanceiraSerializer,
    ContaBancariaSerializer,
    LancamentoSerializer,
    TransferenciaEntreContaSerializer,
)


def _periodo(request):
    hoje = timezone.localdate()
    inicio = request.query_params.get("inicio") or date(hoje.year, hoje.month, 1).isoformat()
    fim = request.query_params.get("fim") or hoje.isoformat()
    return inicio, fim


def _parse_date(value):
    if isinstance(value, date):
        return value
    value = str(value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(value[:10], fmt).date()
        except ValueError:
            continue
    return None


def _sincronizar_recebimento_os(lancamento):
    if (
        lancamento.tipo != Lancamento.Tipo.RECEITA
        or lancamento.status != Lancamento.Status.PAGO
        or not lancamento.os_id
    ):
        return

    OrdemServico.objects.filter(pk=lancamento.os_id).update(
        status_pagamento=OrdemServico.StatusPagamento.PAGO,
        data_recebimento=lancamento.data_pagamento or timezone.localdate(),
        atualizado_em=timezone.now(),
    )


def _parse_decimal(value):
    if isinstance(value, Decimal):
        return value
    value = str(value or "").strip()
    if not value:
        return Decimal("0.00")
    value = value.replace("R$", "").replace(" ", "")
    if "," in value:
        value = value.replace(".", "").replace(",", ".")
    return Decimal(value)


def _normalizar_linha_extrato(row):
    normalizada = {str(key or "").strip().lower(): value for key, value in row.items()}
    data = normalizada.get("data") or normalizada.get("dt") or normalizada.get("date")
    descricao = normalizada.get("descricao") or normalizada.get("descrição") or normalizada.get("historico") or normalizada.get("histórico") or normalizada.get("memo")
    valor = normalizada.get("valor") or normalizada.get("amount") or normalizada.get("vlr")
    documento = normalizada.get("documento") or normalizada.get("doc") or normalizada.get("numero_documento") or normalizada.get("número")
    return {
        "data": _parse_date(data),
        "descricao": str(descricao or "").strip(),
        "valor": _parse_decimal(valor),
        "documento": str(documento or "").strip(),
    }


def _ler_extrato_upload(arquivo):
    nome = arquivo.name.lower()
    if nome.endswith(".xlsx"):
        from openpyxl import load_workbook

        workbook = load_workbook(arquivo, data_only=True, read_only=True)
        sheet = workbook.active
        linhas = list(sheet.iter_rows(values_only=True))
        if not linhas:
            return []
        headers = [str(header or "").strip() for header in linhas[0]]
        return [dict(zip(headers, linha)) for linha in linhas[1:] if any(linha)]

    conteudo = arquivo.read().decode("utf-8-sig")
    amostra = conteudo[:2048]
    try:
        dialect = csv.Sniffer().sniff(amostra, delimiters=";,")
    except csv.Error:
        return list(csv.DictReader(io.StringIO(conteudo), delimiter=";"))
    return list(csv.DictReader(io.StringIO(conteudo), dialect=dialect))


class ContaBancariaViewSet(viewsets.ModelViewSet):
    serializer_class = ContaBancariaSerializer
    filterset_fields = ["ativo", "tipo"]
    search_fields = ["nome", "banco", "agencia", "conta"]
    ordering_fields = ["nome", "saldo_inicial", "criado_em"]

    def get_queryset(self):
        movimento = Sum(
            Case(
                When(lancamentos__tipo=Lancamento.Tipo.RECEITA, lancamentos__status=Lancamento.Status.PAGO, then=F("lancamentos__valor")),
                When(lancamentos__tipo=Lancamento.Tipo.DESPESA, lancamentos__status=Lancamento.Status.PAGO, then=Value(-1) * F("lancamentos__valor")),
                default=Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        return ContaBancaria.objects.annotate(
            saldo_atual=F("saldo_inicial")
            + Coalesce(movimento, Value(Decimal("0.00")), output_field=DecimalField(max_digits=12, decimal_places=2))
        )


class CategoriaFinanceiraViewSet(viewsets.ModelViewSet):
    queryset = CategoriaFinanceira.objects.select_related("pai")
    serializer_class = CategoriaFinanceiraSerializer
    filterset_fields = ["tipo", "pai"]
    search_fields = ["nome"]
    ordering_fields = ["tipo", "nome"]

    @action(detail=False, methods=["get"], url_path="resumo")
    def resumo(self, request):
        inicio, fim = _periodo(request)
        dados = (
            CategoriaFinanceira.objects.filter(lancamentos__data_competencia__range=[inicio, fim])
            .values("id", "nome", "tipo", "cor")
            .annotate(total=Coalesce(Sum("lancamentos__valor"), Decimal("0.00")))
            .order_by("tipo", "nome")
        )
        return Response(list(dados))


class LancamentoViewSet(viewsets.ModelViewSet):
    serializer_class = LancamentoSerializer
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    filterset_fields = ["tipo", "status", "conta_bancaria", "categoria", "os"]
    search_fields = ["descricao", "fornecedor_cliente", "numero_documento", "os__numero"]
    ordering_fields = ["data_vencimento", "data_competencia", "valor", "criado_em"]

    def get_queryset(self):
        queryset = Lancamento.objects.select_related(
            "conta_bancaria",
            "categoria",
            "os",
            "criado_por",
        ).prefetch_related("anexos")
        inicio = self.request.query_params.get("inicio")
        fim = self.request.query_params.get("fim")
        inicio = inicio or self.request.query_params.get("data_vencimento_inicio")
        fim = fim or self.request.query_params.get("data_vencimento_fim")
        if inicio:
            queryset = queryset.filter(data_vencimento__gte=inicio)
        if fim:
            queryset = queryset.filter(data_vencimento__lte=fim)
        return queryset

    def perform_create(self, serializer):
        serializer.save(criado_por=self.request.user)

    @action(detail=True, methods=["post"], url_path="confirmar-pagamento")
    def confirmar_pagamento(self, request, pk=None):
        lancamento = self.get_object()
        lancamento.status = Lancamento.Status.PAGO
        lancamento.data_pagamento = _parse_date(request.data.get("data_pagamento")) or timezone.localdate()
        conta_bancaria = request.data.get("conta_bancaria")
        update_fields = ["status", "data_pagamento"]
        if conta_bancaria:
            lancamento.conta_bancaria_id = conta_bancaria
            update_fields.append("conta_bancaria")
        lancamento.save(update_fields=update_fields)
        _sincronizar_recebimento_os(lancamento)
        for arquivo in request.FILES.getlist("arquivos") or request.FILES.getlist("arquivo"):
            AnexoLancamento.objects.create(
                lancamento=lancamento,
                arquivo=arquivo,
                nome_original=arquivo.name,
            )
        return Response(self.get_serializer(lancamento).data)

    @action(detail=True, methods=["post"], url_path="anexos")
    def anexos(self, request, pk=None):
        lancamento = self.get_object()
        arquivos = request.FILES.getlist("arquivos") or request.FILES.getlist("arquivo")
        if not arquivos:
            return Response({"detail": "Envie ao menos um arquivo."}, status=status.HTTP_400_BAD_REQUEST)
        anexos = [
            AnexoLancamento.objects.create(
                lancamento=lancamento,
                arquivo=arquivo,
                nome_original=arquivo.name,
            )
            for arquivo in arquivos
        ]
        return Response(AnexoLancamentoSerializer(anexos, many=True, context={"request": request}).data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["post"], url_path="importar-extrato")
    def importar_extrato(self, request):
        arquivo = request.FILES.get("arquivo")
        conta_id = request.data.get("conta_bancaria")
        if not arquivo:
            return Response({"detail": "Envie o arquivo do extrato."}, status=status.HTTP_400_BAD_REQUEST)
        if not conta_id:
            return Response({"detail": "Selecione a conta bancaria do extrato."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            conta = ContaBancaria.objects.get(pk=conta_id)
        except ContaBancaria.DoesNotExist:
            return Response({"detail": "Conta bancaria nao encontrada."}, status=status.HTTP_400_BAD_REQUEST)
        conciliados = []
        criados = []
        ignorados = []

        for index, row in enumerate(_ler_extrato_upload(arquivo), start=2):
            item = _normalizar_linha_extrato(row)
            if not item["data"] or not item["descricao"] or item["valor"] == Decimal("0.00"):
                ignorados.append({"linha": index, "motivo": "Linha sem data, descricao ou valor."})
                continue

            tipo = Lancamento.Tipo.RECEITA if item["valor"] > 0 else Lancamento.Tipo.DESPESA
            valor = abs(item["valor"]).quantize(Decimal("0.01"))
            data_inicio = item["data"] - timedelta(days=7)
            data_fim = item["data"] + timedelta(days=7)
            candidatos = Lancamento.objects.filter(
                tipo=tipo,
                valor=valor,
                status__in=[Lancamento.Status.PENDENTE, Lancamento.Status.ATRASADO],
                data_vencimento__range=[data_inicio, data_fim],
            ).order_by("data_vencimento", "id")
            if item["documento"]:
                por_documento = candidatos.filter(numero_documento__icontains=item["documento"]).first()
                lancamento = por_documento or candidatos.first()
            else:
                lancamento = candidatos.first()

            if lancamento:
                observacao = f"Conciliado automaticamente pelo extrato em {timezone.localdate().isoformat()}."
                lancamento.status = Lancamento.Status.PAGO
                lancamento.data_pagamento = item["data"]
                lancamento.conta_bancaria = conta
                lancamento.observacoes = f"{lancamento.observacoes}\n{observacao}".strip()
                lancamento.save(update_fields=["status", "data_pagamento", "conta_bancaria", "observacoes"])
                _sincronizar_recebimento_os(lancamento)
                conciliados.append({"linha": index, "lancamento": lancamento.id, "descricao": lancamento.descricao})
                continue

            categoria, _ = CategoriaFinanceira.objects.get_or_create(
                nome="Receitas conciliadas" if tipo == Lancamento.Tipo.RECEITA else "Despesas conciliadas",
                tipo=tipo,
                defaults={"cor": "#10B981" if tipo == Lancamento.Tipo.RECEITA else "#EF4444"},
            )
            lancamento = Lancamento.objects.create(
                tipo=tipo,
                descricao=item["descricao"][:255],
                valor=valor,
                data_competencia=item["data"],
                data_vencimento=item["data"],
                data_pagamento=item["data"],
                status=Lancamento.Status.PAGO,
                conta_bancaria=conta,
                categoria=categoria,
                numero_documento=item["documento"],
                observacoes="Criado automaticamente pela importacao do extrato bancario.",
                criado_por=request.user,
            )
            criados.append({"linha": index, "lancamento": lancamento.id, "descricao": lancamento.descricao})

        return Response(
            {
                "conciliados": len(conciliados),
                "criados": len(criados),
                "ignorados": len(ignorados),
                "itens_conciliados": conciliados,
                "itens_criados": criados,
                "itens_ignorados": ignorados,
            }
        )


class TransferenciaEntreContaViewSet(viewsets.ModelViewSet):
    serializer_class = TransferenciaEntreContaSerializer
    filterset_fields = ["conta_origem", "conta_destino"]
    search_fields = ["descricao", "conta_origem__nome", "conta_destino__nome"]
    ordering_fields = ["data", "valor", "criado_em"]

    def get_queryset(self):
        return TransferenciaEntreConta.objects.select_related(
            "conta_origem",
            "conta_destino",
            "criado_por",
        )

    def perform_create(self, serializer):
        serializer.save(criado_por=self.request.user)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def dashboard(request):
    inicio, fim = _periodo(request)
    hoje = timezone.localdate()
    lancamentos = Lancamento.objects.filter(data_competencia__range=[inicio, fim])
    receita = lancamentos.filter(tipo=Lancamento.Tipo.RECEITA).exclude(status=Lancamento.Status.CANCELADO).aggregate(total=Coalesce(Sum("valor"), Decimal("0.00")))["total"]
    despesa = lancamentos.filter(tipo=Lancamento.Tipo.DESPESA).exclude(status=Lancamento.Status.CANCELADO).aggregate(total=Coalesce(Sum("valor"), Decimal("0.00")))["total"]
    receber = Lancamento.objects.filter(tipo=Lancamento.Tipo.RECEITA, status__in=[Lancamento.Status.PENDENTE, Lancamento.Status.ATRASADO]).aggregate(total=Coalesce(Sum("valor"), Decimal("0.00")))["total"]
    pagar = Lancamento.objects.filter(tipo=Lancamento.Tipo.DESPESA, status__in=[Lancamento.Status.PENDENTE, Lancamento.Status.ATRASADO]).aggregate(total=Coalesce(Sum("valor"), Decimal("0.00")))["total"]
    saldo_total = ContaBancaria.objects.aggregate(total=Coalesce(Sum("saldo_inicial"), Decimal("0.00")))["total"]

    por_mes = (
        Lancamento.objects.exclude(status=Lancamento.Status.CANCELADO)
        .annotate(mes=TruncMonth("data_competencia"))
        .values("mes", "tipo")
        .annotate(total=Coalesce(Sum("valor"), Decimal("0.00")))
        .order_by("mes")
    )
    despesas_categoria = (
        Lancamento.objects.filter(tipo=Lancamento.Tipo.DESPESA, data_competencia__range=[inicio, fim])
        .exclude(status=Lancamento.Status.CANCELADO)
        .values("categoria__nome", "categoria__cor")
        .annotate(total=Coalesce(Sum("valor"), Decimal("0.00")))
        .order_by("-total")
    )
    contas_receber = LancamentoSerializer(Lancamento.objects.filter(tipo=Lancamento.Tipo.RECEITA, status__in=[Lancamento.Status.PENDENTE, Lancamento.Status.ATRASADO]).order_by("data_vencimento")[:10], many=True).data
    contas_pagar = LancamentoSerializer(Lancamento.objects.filter(tipo=Lancamento.Tipo.DESPESA, status__in=[Lancamento.Status.PENDENTE, Lancamento.Status.ATRASADO]).order_by("data_vencimento")[:10], many=True).data
    receber_atrasado_qs = Lancamento.objects.filter(
        tipo=Lancamento.Tipo.RECEITA,
        status__in=[Lancamento.Status.PENDENTE, Lancamento.Status.ATRASADO],
        data_vencimento__lt=hoje,
    )
    pagar_atrasado_qs = Lancamento.objects.filter(
        tipo=Lancamento.Tipo.DESPESA,
        status__in=[Lancamento.Status.PENDENTE, Lancamento.Status.ATRASADO],
        data_vencimento__lt=hoje,
    )

    return Response(
        {
            "receita": receita,
            "despesa": despesa,
            "lucro": receita - despesa,
            "contas_receber": receber,
            "contas_pagar": pagar,
            "saldo_total": saldo_total,
            "por_mes": list(por_mes),
            "despesas_categoria": list(despesas_categoria),
            "contas_receber_lista": contas_receber,
            "contas_pagar_lista": contas_pagar,
            "receber_atrasado": receber_atrasado_qs.aggregate(
                total=Coalesce(Sum("valor"), Decimal("0.00"))
            )["total"],
            "receber_atrasado_count": receber_atrasado_qs.count(),
            "pagar_atrasado": pagar_atrasado_qs.aggregate(
                total=Coalesce(Sum("valor"), Decimal("0.00"))
            )["total"],
            "pagar_atrasado_count": pagar_atrasado_qs.count(),
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def fluxo_caixa(request):
    inicio, fim = _periodo(request)
    dados = (
        Lancamento.objects.filter(data_vencimento__range=[inicio, fim])
        .exclude(status=Lancamento.Status.CANCELADO)
        .values("data_vencimento", "tipo")
        .annotate(total=Coalesce(Sum("valor"), Decimal("0.00")))
        .order_by("data_vencimento")
    )
    return Response(list(dados))


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def relatorio_dre(request):
    inicio, fim = _periodo(request)
    base = Lancamento.objects.filter(data_competencia__range=[inicio, fim]).exclude(status=Lancamento.Status.CANCELADO)
    receita = base.filter(tipo=Lancamento.Tipo.RECEITA).aggregate(total=Coalesce(Sum("valor"), Decimal("0.00")))["total"]
    despesa = base.filter(tipo=Lancamento.Tipo.DESPESA).aggregate(total=Coalesce(Sum("valor"), Decimal("0.00")))["total"]
    os_periodo = OrdemServico.objects.filter(criado_em__date__range=[inicio, fim]).values("status").annotate(total=Coalesce(Sum("valor_total_orcado"), Decimal("0.00")))
    hoje = timezone.localdate()
    aging = {
        "a_vencer": Lancamento.objects.filter(tipo=Lancamento.Tipo.RECEITA, status=Lancamento.Status.PENDENTE, data_vencimento__gte=hoje).aggregate(total=Coalesce(Sum("valor"), Decimal("0.00")))["total"],
        "ate_30": Lancamento.objects.filter(tipo=Lancamento.Tipo.RECEITA, status=Lancamento.Status.ATRASADO, data_vencimento__gte=hoje.replace(day=1)).aggregate(total=Coalesce(Sum("valor"), Decimal("0.00")))["total"],
        "vencidos": Lancamento.objects.filter(tipo=Lancamento.Tipo.RECEITA, status=Lancamento.Status.ATRASADO).aggregate(total=Coalesce(Sum("valor"), Decimal("0.00")))["total"],
    }
    return Response(
        {
            "receita_bruta": receita,
            "despesas": despesa,
            "resultado": receita - despesa,
            "margem": (receita - despesa) / receita * 100 if receita else Decimal("0.00"),
            "ordens_servico": list(os_periodo),
            "aging_receber": aging,
        }
    )
